"""Google Drive budget scanner — extracts structured budget data from financial
documents on shared drives and associates them with project numbers.

Produces:
  knowledge/budgets/{project_code}.md  — per-project budget knowledge
  knowledge/budgets/summary.md         — cross-project financial overview

Handles:
  - Google Docs/Sheets (exported via Drive API)
  - PDFs (downloaded + pypdf)
  - DOCX files (downloaded + python-docx)
  - XLSX files (downloaded + openpyxl)

Modes:
  full        — all matching documents (no time filter)
  1yr         — documents modified in the last year
  incremental — documents modified since last scan
"""

import io
import os
import re
import time
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

from google.oauth2 import service_account
from googleapiclient.discovery import build

from .base import (
    KNOWLEDGE_DIR,
    distill_to_markdown,
    get_last_scan,
    update_scan_timestamp,
    write_knowledge_file,
    get_claude_client,
    DISTILL_MODEL,
)

BUDGET_KNOWLEDGE_DIR = KNOWLEDGE_DIR / "budgets"

# Rate limiting
PAGE_LIST_DELAY = 0.1
EXPORT_DELAY = 0.5
MAX_DOC_CHARS = 60000

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

# MIME types
GOOGLE_DOC_MIME = "application/vnd.google-apps.document"
GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
PDF_MIME = "application/pdf"
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Keywords that indicate a budget/financial document
BUDGET_KEYWORDS = [
    "budget", "invoice", "cost", "pricing", "quote", "funding",
    "financial", "expenditure", "expense", "billing", "payment",
    "clin", "contract line", "obligation", "disbursement",
    "cost proposal", "cost volume", "burn rate",
    "purchase order", "po ", "p.o.",
]

# Folder keywords
BUDGET_FOLDER_KEYWORDS = [
    "budget", "invoice", "financial", "billing", "cost",
    "accounting", "payments",
]

# Regex to extract project codes like [200-11] or 200-11 from paths
PROJECT_CODE_RE = re.compile(r'\[?(\d{3}-\d{1,2})\]?')


def _get_credentials(user_email):
    import json
    import base64
    key_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    if not key_json:
        return None
    key_data = json.loads(base64.b64decode(key_json))
    creds = service_account.Credentials.from_service_account_info(key_data, scopes=SCOPES)
    return creds.with_subject(user_email)


def _get_drive_service(user_email):
    creds = _get_credentials(user_email)
    if not creds:
        raise RuntimeError("No Google service account credentials configured")
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _get_shared_drive_ids(service):
    target_drives = os.environ.get("GOOGLE_SHARED_DRIVES", "Sales,Federal Projects")
    target_names = [n.strip().lower() for n in target_drives.split(",")]
    drive_ids = []
    try:
        results = service.drives().list(pageSize=50).execute()
        for drive in results.get("drives", []):
            if drive["name"].lower() in target_names:
                drive_ids.append({"id": drive["id"], "name": drive["name"]})
    except Exception as e:
        print(f"  [ERROR] Could not list shared drives: {e}")
    return drive_ids


def _is_budget_doc(file_name, parent_path=""):
    """Check if a file looks like a budget/financial document."""
    combined = f"{parent_path}/{file_name}".lower()
    return any(kw in combined for kw in BUDGET_KEYWORDS + BUDGET_FOLDER_KEYWORDS)


def _extract_project_code(file_name, parent_path=""):
    """Try to extract a BST project code (e.g., 200-11) from file/folder path."""
    combined = f"{parent_path}/{file_name}"
    match = PROJECT_CODE_RE.search(combined)
    return match.group(1) if match else None


def _fetch_folder_map(service, drive_id):
    folder_map = {}
    page_token = None
    while True:
        params = {
            "q": "trashed = false and mimeType = 'application/vnd.google-apps.folder'",
            "fields": "nextPageToken,files(id,name,parents)",
            "pageSize": 1000,
            "corpora": "drive",
            "driveId": drive_id,
            "includeItemsFromAllDrives": True,
            "supportsAllDrives": True,
        }
        if page_token:
            params["pageToken"] = page_token
        result = service.files().list(**params).execute()
        for folder in result.get("files", []):
            folder_map[folder["id"]] = {
                "name": folder["name"],
                "parent": (folder.get("parents") or [None])[0],
            }
        page_token = result.get("nextPageToken")
        if not page_token:
            break
        time.sleep(PAGE_LIST_DELAY)
    return folder_map


def _resolve_parent_path(file_dict, folder_map, max_depth=10):
    parts = []
    parent_id = (file_dict.get("parents") or [None])[0]
    depth = 0
    while parent_id and parent_id in folder_map and depth < max_depth:
        folder = folder_map[parent_id]
        parts.append(folder["name"])
        parent_id = folder["parent"]
        depth += 1
    parts.reverse()
    return "/".join(parts)


def _list_budget_files(service, drive_id, time_filter=None):
    """List all budget/financial files from a shared drive."""
    target_mimes = [GOOGLE_DOC_MIME, GOOGLE_SHEET_MIME, PDF_MIME, DOCX_MIME, XLSX_MIME]
    mime_clauses = " or ".join(f"mimeType = '{m}'" for m in target_mimes)
    query_parts = ["trashed = false", f"({mime_clauses})"]
    if time_filter:
        query_parts.append(f"modifiedTime > '{time_filter}'")
    query = " and ".join(query_parts)

    file_fields = (
        "id,name,mimeType,modifiedTime,createdTime,webViewLink,"
        "parents,owners,lastModifyingUser"
    )

    all_files = []
    page_token = None
    while True:
        params = {
            "q": query,
            "fields": f"nextPageToken,files({file_fields})",
            "orderBy": "modifiedTime desc",
            "pageSize": 1000,
            "corpora": "drive",
            "driveId": drive_id,
            "includeItemsFromAllDrives": True,
            "supportsAllDrives": True,
        }
        if page_token:
            params["pageToken"] = page_token
        result = service.files().list(**params).execute()
        all_files.extend(result.get("files", []))
        page_token = result.get("nextPageToken")
        if not page_token:
            break
        time.sleep(PAGE_LIST_DELAY)

    folder_map = _fetch_folder_map(service, drive_id)

    candidates = []
    for f in all_files:
        parent_path = _resolve_parent_path(f, folder_map)
        if _is_budget_doc(f["name"], parent_path):
            f["_parent_path"] = parent_path
            f["_project_code"] = _extract_project_code(f["name"], parent_path)
            candidates.append(f)

    return candidates


# --- Text extraction (shared with proposals scanner) ---

def _extract_google_doc(service, file_id):
    try:
        content = service.files().export(fileId=file_id, mimeType="text/plain").execute()
        text = content.decode("utf-8") if isinstance(content, bytes) else str(content)
        return text[:MAX_DOC_CHARS]
    except Exception as e:
        print(f"    [WARN] Could not export Google Doc {file_id}: {e}")
        return None


def _extract_google_sheet(service, file_id):
    try:
        content = service.files().export(fileId=file_id, mimeType="text/csv").execute()
        text = content.decode("utf-8") if isinstance(content, bytes) else str(content)
        return text[:MAX_DOC_CHARS]
    except Exception as e:
        print(f"    [WARN] Could not export Google Sheet {file_id}: {e}")
        return None


def _download_file_bytes(service, file_id):
    try:
        return service.files().get_media(fileId=file_id).execute()
    except Exception as e:
        print(f"    [WARN] Could not download file {file_id}: {e}")
        return None


def _extract_pdf_text(file_bytes):
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(file_bytes))
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
            if sum(len(p) for p in pages) > MAX_DOC_CHARS:
                break
        return "\n\n".join(pages)[:MAX_DOC_CHARS]
    except ImportError:
        print("    [WARN] pypdf not installed — pip install pypdf")
        return None
    except Exception as e:
        print(f"    [WARN] PDF extraction failed: {e}")
        return None


def _extract_docx_text(file_bytes):
    try:
        from docx import Document
        doc = Document(io.BytesIO(file_bytes))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())[:MAX_DOC_CHARS]
    except ImportError:
        print("    [WARN] python-docx not installed — pip install python-docx")
        return None
    except Exception as e:
        print(f"    [WARN] DOCX extraction failed: {e}")
        return None


def _extract_xlsx_text(file_bytes):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"=== Sheet: {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append(",".join(cells))
            if sum(len(l) for l in lines) > MAX_DOC_CHARS:
                break
        wb.close()
        return "\n".join(lines)[:MAX_DOC_CHARS]
    except ImportError:
        print("    [WARN] openpyxl not installed — pip install openpyxl")
        return None
    except Exception as e:
        print(f"    [WARN] XLSX extraction failed: {e}")
        return None


def _extract_text(service, file_dict):
    mime = file_dict["mimeType"]
    file_id = file_dict["id"]

    if mime == GOOGLE_DOC_MIME:
        return _extract_google_doc(service, file_id)
    elif mime == GOOGLE_SHEET_MIME:
        return _extract_google_sheet(service, file_id)
    elif mime in (PDF_MIME, DOCX_MIME, XLSX_MIME):
        file_bytes = _download_file_bytes(service, file_id)
        if not file_bytes:
            return None
        if mime == PDF_MIME:
            return _extract_pdf_text(file_bytes)
        elif mime == DOCX_MIME:
            return _extract_docx_text(file_bytes)
        else:
            return _extract_xlsx_text(file_bytes)
    return None


def _safe_filename(name):
    name = re.sub(r'[^\w\s-]', '', name)
    name = re.sub(r'\s+', '_', name.strip())
    return name[:80].lower()


BUDGET_DISTILL_PROMPT = """\
You are distilling budget and financial documents for a Black Swift Technologies (BST) project \
into a structured knowledge file. This will be used to answer questions about project budgets, \
spending, and financial status.

Output a markdown file:

# Project {Project Code} — Budget & Financials

## Budget Overview
- Contract/award value (total funded amount)
- Period of performance
- Funding agency / client
- Contract type (SBIR Phase I/II, IRAD, commercial, etc.)

## Budget Breakdown
Extract line items, CLINs, or budget categories:
- Labor (by person or role if available)
- Travel
- Equipment / materials
- Subcontracts
- Other direct costs
- Indirect costs / overhead / G&A
- Fee / profit

## Invoices & Billing
List invoices found with: invoice number, date, amount, period covered, status if known.

## Purchase Orders
List POs with: PO number, vendor, amount, description, date.

## Funding Modifications
Any contract mods, additional funding, de-obligations, or budget reallocations.

## Key Financial Details
- Burn rate observations
- Budget vs actual comparisons if present
- Remaining balance if calculable
- Any cost overruns or underspends noted

Be precise with dollar amounts, dates, and document references. \
If multiple documents cover the same project, merge the information. \
Note which source document each data point comes from."""


BUDGET_SUMMARY_PROMPT = """\
You are creating a financial overview across all BST projects. \
Summarize the budget data into a reference document:

# BST Project Financial Overview

## Active Project Budgets
For each project, show: project code, client, total budget, contract type, \
and any spend/remaining data if known.

## By Funding Source
Group projects by agency/client with total funded amounts.

## By Contract Type
Group by SBIR Phase I, SBIR Phase II, IRAD, commercial, etc.

## Financial Health Indicators
- Projects with known budget concerns
- Projects nearing end of funding
- Large upcoming invoices or milestones

## QuickBooks Integration Notes
Note which projects have matching QuickBooks data (if referenced in docs) \
and which are Drive-only.

Be factual and include dollar amounts where available."""


def scan_all(mode="full", delegate_email=None, progress_callback=None):
    """Scan budget/financial documents from shared drives, grouped by project.

    Returns list of (project_code_or_name, doc_count, output_path) tuples.
    """
    if delegate_email is None:
        delegate_email = os.environ.get(
            "GOOGLE_DELEGATE_EMAIL",
            "jack.elston@blackswifttech.com",
        )

    service = _get_drive_service(delegate_email)
    shared_drives = _get_shared_drive_ids(service)

    if not shared_drives:
        print("[ERROR] No matching shared drives found")
        return []

    print(f"Found {len(shared_drives)} shared drives")

    # Time filter
    time_filter = None
    if mode == "1yr":
        one_year_ago = datetime.now(timezone.utc) - timedelta(days=365)
        time_filter = one_year_ago.isoformat()
        print(f"1yr mode: docs modified since {time_filter[:10]}")
    elif mode == "incremental":
        last = get_last_scan("budgets")
        if last:
            time_filter = last.isoformat()
            print(f"Incremental: changes since {time_filter}")
        else:
            print("No previous scan — falling back to full")
            mode = "full"

    # Collect all budget files across drives
    all_candidates = []
    for drive_info in shared_drives:
        print(f"\n  Searching {drive_info['name']} for budget/financial docs...")
        candidates = _list_budget_files(service, drive_info["id"], time_filter=time_filter)
        print(f"  Found {len(candidates)} budget documents in {drive_info['name']}")
        all_candidates.extend(candidates)

    if not all_candidates:
        print("No budget documents found")
        update_scan_timestamp("budgets")
        return []

    print(f"\nTotal budget documents: {len(all_candidates)}")

    # Group by project code
    by_project = defaultdict(list)
    unassigned = []
    for f in all_candidates:
        code = f.get("_project_code")
        if code:
            by_project[code].append(f)
        else:
            unassigned.append(f)

    print(f"Grouped into {len(by_project)} projects + {len(unassigned)} unassigned docs")

    # Process each project group
    results = []
    project_codes = sorted(by_project.keys())

    for i, code in enumerate(project_codes):
        files = by_project[code]
        print(f"\n[{i + 1}/{len(project_codes)}] Project {code} ({len(files)} docs)")

        # Skip if output file already exists (resume support)
        filename = f"project_{code.replace('-', '_')}.md"
        output_path = BUDGET_KNOWLEDGE_DIR / filename
        if output_path.exists() and output_path.stat().st_size > 100:
            print(f"  [SKIP] Already scanned — {filename}")
            results.append((code, len(files), output_path))
            continue

        # Extract text from all docs for this project
        doc_texts = []
        for f in files:
            text = _extract_text(service, f)
            if text and len(text.strip()) > 50:
                owner = (f.get("lastModifyingUser") or {}).get("displayName", "")
                mod = f.get("modifiedTime", "")[:10]
                header = (
                    f"\n--- Document: {f['name']} ---\n"
                    f"Path: {f.get('_parent_path', '')}/{f['name']}\n"
                    f"Modified: {mod} | Editor: {owner}\n"
                    f"Link: {f.get('webViewLink', '')}\n\n"
                )
                doc_texts.append(header + text)
            time.sleep(EXPORT_DELAY)

        if not doc_texts:
            print(f"  [SKIP] Project {code}: no extractable text")
            continue

        raw_input = (
            f"PROJECT CODE: {code}\n"
            f"Number of financial documents: {len(doc_texts)}\n\n"
            + "\n".join(doc_texts)
        )

        filename = f"project_{code.replace('-', '_')}.md"
        output_path = BUDGET_KNOWLEDGE_DIR / filename

        print(f"  Distilling {code} ({len(doc_texts)} docs, {sum(len(t) for t in doc_texts)} chars)...")
        result = distill_to_markdown(raw_input, BUDGET_DISTILL_PROMPT, output_path, max_tokens=3000)

        if result:
            results.append((code, len(doc_texts), output_path))

    # Handle unassigned docs (no project code found) — group into one file
    unassigned_path = BUDGET_KNOWLEDGE_DIR / "unassigned.md"
    if unassigned and not (unassigned_path.exists() and unassigned_path.stat().st_size > 100):
        # Cap at 30 unassigned docs to avoid getting stuck on large collections
        unassigned_capped = unassigned[:30]
        print(f"\nProcessing {len(unassigned_capped)} of {len(unassigned)} unassigned budget docs...")
        doc_texts = []
        for f in unassigned_capped:
            text = _extract_text(service, f)
            if text and len(text.strip()) > 50:
                header = (
                    f"\n--- Document: {f['name']} ---\n"
                    f"Path: {f.get('_parent_path', '')}/{f['name']}\n"
                    f"Modified: {f.get('modifiedTime', '')[:10]}\n\n"
                )
                doc_texts.append(header + text)
            time.sleep(EXPORT_DELAY)
    elif unassigned and unassigned_path.exists():
        print(f"\n[SKIP] Unassigned docs already scanned — unassigned.md")
        results.append(("unassigned", len(unassigned), unassigned_path))
        doc_texts = []  # skip the distillation block below

        if doc_texts:
            raw_input = (
                f"UNASSIGNED BUDGET DOCUMENTS (no project code detected)\n"
                f"Number of documents: {len(doc_texts)}\n\n"
                + "\n".join(doc_texts)
            )
            output_path = BUDGET_KNOWLEDGE_DIR / "unassigned.md"
            distill_to_markdown(raw_input, BUDGET_DISTILL_PROMPT, output_path, max_tokens=3000)
            results.append(("unassigned", len(doc_texts), output_path))

    # Generate summary
    if results:
        _generate_summary(results)

    update_scan_timestamp("budgets")
    return results


def _generate_summary(results):
    """Generate cross-project financial summary."""
    snippets = []
    for code, doc_count, path in results:
        if path.exists():
            content = path.read_text()
            snippets.append(f"### Project {code} ({doc_count} docs)\n{content[:800]}")

    if not snippets:
        return

    combined = "\n\n".join(snippets)
    if len(combined) > 90000:
        combined = combined[:90000] + "\n\n[TRUNCATED]"

    client = get_claude_client()
    for attempt in range(5):
        try:
            message = client.messages.create(
                model=DISTILL_MODEL,
                max_tokens=3000,
                system=BUDGET_SUMMARY_PROMPT,
                messages=[{"role": "user", "content": f"Project budget data:\n\n{combined}"}],
            )
            summary_path = BUDGET_KNOWLEDGE_DIR / "summary.md"
            write_knowledge_file(summary_path, message.content[0].text)
            print(f"  Budget summary written to {summary_path}")
            return
        except Exception as e:
            if "rate_limit" in str(e).lower():
                wait = 30 * (attempt + 1)
                print(f"  Rate limited, waiting {wait}s...")
                time.sleep(wait)
            else:
                print(f"  [WARN] Summary generation failed: {e}")
                return
